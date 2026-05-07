"""
Score composite "{color} curtain" phrases against "{ambiance} living room" queries.

For every color atom (23) we build a composite item text — e.g. "deep blue curtain",
"warm white curtain" — embed it, and score it against all 400 combined (space, ambiance)
queries using the project's standard 0.25/0.25/0.50 weighted scoring. We then enrich
with discriminative & z-scores (per-row across all 400 queries, consistent with
atom_scores.csv), filter to the 20 living-room queries, and pivot.

Outputs:
  - data/processed/scores_color_curtain_pairs_living_room_zscore.csv
  - data/processed/scores_color_curtain_pairs_living_room_zscore.xlsx (heatmap)
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from semantic_architecture.app_state import AppState
from semantic_architecture.io_utils import save_csv
from semantic_architecture.scoring import (
    ScoringWeights,
    enrich_with_discriminative_scores,
    score_items_against_queries,
)


SPACE = "living room"
SCORE_COL = "zscore_score"
COLLECTION = "color_curtain_composites"
OUT_CSV = "data/processed/scores_color_curtain_pairs_living_room_zscore.csv"
OUT_XLSX = "data/processed/scores_color_curtain_pairs_living_room_zscore.xlsx"


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    root = Path(__file__).resolve().parent.parent
    state = AppState.load(config_path=root / "config/config.yaml")

    colors = [a for a in state.atoms if a.family == "color"]
    if not colors:
        print("[error] No color atoms found.", file=sys.stderr)
        return 1

    item_texts = [f"{c.text} curtain" for c in colors]
    item_ids = [f"comp_{c.id}_curtain" for c in colors]
    item_families = ["composite_color_curtain"] * len(colors)

    from semantic_architecture.embeddings import EmbeddingModel
    emb_cfg = state.config.get("embedding", {})
    model = EmbeddingModel(
        model_name=emb_cfg.get("model_name", "all-MiniLM-L6-v2"),
        cache_dir=root / emb_cfg.get("cache_dir", "data/processed/embeddings"),
        batch_size=emb_cfg.get("batch_size", 64),
    )
    item_embs = model.load_or_compute(item_texts, COLLECTION)

    weights = ScoringWeights.from_config(state.config.get("scoring", {}))
    scores = score_items_against_queries(
        item_texts=item_texts,
        item_families=item_families,
        item_ids=item_ids,
        item_embeddings=item_embs,
        space_queries=state.space_queries,
        ambiance_queries=state.ambiance_queries,
        combined_queries=state.combined_queries,
        space_embeddings=state.space_embeddings,
        ambiance_embeddings=state.ambiance_embeddings,
        combined_embeddings=state.combined_embeddings,
        weights=weights,
    )
    scores = enrich_with_discriminative_scores(scores)

    sub = scores.loc[
        scores["space_text"] == SPACE,
        ["text", "ambiance_text", SCORE_COL, "weighted_score"],
    ]
    if sub.empty:
        print(f"[error] No rows for space={SPACE!r}.", file=sys.stderr)
        return 1

    pivot_z = sub.pivot(index="text", columns="ambiance_text", values=SCORE_COL)
    pivot_w = sub.pivot(index="text", columns="ambiance_text", values="weighted_score")

    amb_order = [a["text"] for a in state.ambiances if a["text"] in pivot_z.columns]
    pivot_z = pivot_z[amb_order]
    pivot_w = pivot_w[amb_order]

    row_order = sorted(pivot_z.index)
    pivot_z = pivot_z.loc[row_order]
    pivot_w = pivot_w.loc[row_order]

    save_csv(
        pivot_z,
        root / OUT_CSV,
        desc=f"{SCORE_COL} pivot ({{color}} curtain × {{ambiance}} {SPACE})",
    )

    write_xlsx(pivot_z, pivot_w, root / OUT_XLSX)

    with pd.option_context(
        "display.max_columns", None,
        "display.width", 240,
        "display.float_format", lambda v: f"{v: .3f}",
    ):
        print(
            f"\n{SCORE_COL} — '{{color}} curtain' × '{{ambiance}} {SPACE}'  "
            f"(shape: {pivot_z.shape[0]}×{pivot_z.shape[1]})\n"
        )
        print(pivot_z.to_string())

    return 0


def write_xlsx(pivot_z: pd.DataFrame, pivot_w: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    title_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    row_label_font = Font(name="Arial", size=10, bold=True)

    header_fill = PatternFill("solid", start_color="2F4F4F")
    row_label_fill = PatternFill("solid", start_color="EEEEEE")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_sheet(name: str, df: pd.DataFrame, subtitle: str,
                  scale_min: float, scale_max: float, fmt: str) -> None:
        ws = wb.create_sheet(name)
        n_rows, n_cols = df.shape

        title = (
            f"{subtitle}  —  '{{color}} curtain'  ×  '{{ambiance}} {SPACE}'"
        )
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols + 1)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(row=3, column=1, value="color × curtain  \\  ambiance × living room").font = header_font
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=3, column=1).border = border

        for j, col in enumerate(df.columns, start=2):
            c = ws.cell(row=3, column=j, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

        for i, idx in enumerate(df.index, start=4):
            label = ws.cell(row=i, column=1, value=idx)
            label.font = row_label_font
            label.fill = row_label_fill
            label.alignment = Alignment(horizontal="left", vertical="center")
            label.border = border

            for j, col in enumerate(df.columns, start=2):
                v = df.iloc[i - 4, j - 2]
                cell = ws.cell(row=i, column=j, value=float(v) if pd.notna(v) else None)
                cell.font = body_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = fmt
                cell.border = border

        last_col_letter = get_column_letter(n_cols + 1)
        data_range = f"B4:{last_col_letter}{3 + n_rows}"
        rule = ColorScaleRule(
            start_type="num", start_value=scale_min, start_color="2C7BB6",
            mid_type="num", mid_value=(scale_min + scale_max) / 2, mid_color="FFFFFF",
            end_type="num", end_value=scale_max, end_color="D7191C",
        )
        ws.conditional_formatting.add(data_range, rule)

        ws.column_dimensions["A"].width = 26
        for j in range(2, n_cols + 2):
            ws.column_dimensions[get_column_letter(j)].width = 11
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[3].height = 32
        ws.freeze_panes = "B4"
        ws.sheet_view.zoomScale = 110

        legend_row = 4 + n_rows + 1
        ws.cell(row=legend_row, column=1, value="Notes").font = title_font
        notes = [
            "Items are composite phrases '{color} curtain' (23 rows). Queries are '{ambiance} living room' (20 cols).",
            "Each cell uses the standard pipeline scoring: 0.25·sim_space + 0.25·sim_ambiance + 0.50·sim_combined.",
            "z-score = (weighted_score − row mean across ALL 400 queries) / row std. So this 20-column slice can be skewed per row.",
            "Positive (red) → composite is more aligned with this ambiance than its own average; negative (blue) → less.",
        ]
        for k, line in enumerate(notes):
            ws.cell(row=legend_row + 1 + k, column=1, value=line).font = body_font
            ws.merge_cells(
                start_row=legend_row + 1 + k, start_column=1,
                end_row=legend_row + 1 + k, end_column=min(n_cols + 1, 10),
            )

    add_sheet("z-score", pivot_z, "Z-score (per-row, normalised)",
              scale_min=-2.5, scale_max=2.5, fmt="0.000;-0.000;-")

    w_min = float(pivot_w.values.min())
    w_max = float(pivot_w.values.max())
    add_sheet("weighted_score", pivot_w, "Weighted score (raw)",
              scale_min=w_min, scale_max=w_max, fmt="0.0000")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[io] Saved styled XLSX → {path}  (sheets: z-score, weighted_score)")


if __name__ == "__main__":
    sys.exit(main())
