"""
Pivot pre-computed atom scores for: all colors + Curtain × all ambiances × living room.

Reads data/processed/atom_scores.csv (built by `python -m semantic_architecture.cli export`),
filters to the 23 color atoms plus dec_curtain inside the "living room" space, and pivots
to a 24 x 20 table of zscore_score values (rows = atom, cols = ambiance, in source order).

Outputs:
  - data/processed/scores_colors_curtain_living_room_zscore.csv  (raw)
  - data/processed/scores_colors_curtain_living_room_zscore.xlsx (styled heatmap)
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from semantic_architecture.io_utils import load_csv, save_csv
from semantic_architecture.queries import load_ambiances


SPACE = "living room"
CURTAIN_ID = "dec_curtain"
SCORE_COL = "zscore_score"
OUT_REL = "data/processed/scores_colors_curtain_living_room_zscore.csv"
XLSX_REL = "data/processed/scores_colors_curtain_living_room_zscore.xlsx"


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    root = Path(__file__).resolve().parent.parent
    scores_path = root / "data/processed/atom_scores.csv"
    if not scores_path.exists():
        print(
            f"[error] {scores_path} not found. "
            "Run `python -m semantic_architecture.cli export` first.",
            file=sys.stderr,
        )
        return 1

    df = load_csv(scores_path)

    mask = (
        (df["space_text"] == SPACE)
        & ((df["family"] == "color") | (df["item_id"] == CURTAIN_ID))
    )
    sub = df.loc[mask, ["text", "item_id", "family", "ambiance_text", SCORE_COL]]

    if sub.empty:
        print(f"[error] No rows matched space={SPACE!r}, color | {CURTAIN_ID}.", file=sys.stderr)
        return 1

    pivot = sub.pivot(index="text", columns="ambiance_text", values=SCORE_COL)

    ambiances = load_ambiances(root / "data/raw/ambiances.json")
    amb_order = [a["text"] for a in ambiances if a["text"] in pivot.columns]
    pivot = pivot[amb_order]

    color_rows = sorted(t for t in pivot.index if t != "curtain")
    row_order = color_rows + [t for t in pivot.index if t == "curtain"]
    pivot = pivot.loc[row_order]

    save_csv(pivot, root / OUT_REL, desc=f"{SCORE_COL} pivot (colors + curtain × ambiances @ {SPACE})")

    write_xlsx(pivot, root / XLSX_REL)

    with pd.option_context(
        "display.max_columns", None,
        "display.width", 220,
        "display.float_format", lambda v: f"{v: .3f}",
    ):
        print(f"\n{SCORE_COL} — colors + curtain × ambiances @ '{SPACE}'  "
              f"(shape: {pivot.shape[0]}×{pivot.shape[1]})\n")
        print(pivot.to_string())

    return 0


def write_xlsx(pivot: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "z-score"

    title_font = Font(name="Arial", size=12, bold=True)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    row_label_font = Font(name="Arial", size=10, bold=True)
    curtain_font = Font(name="Arial", size=10, bold=True, italic=True)

    header_fill = PatternFill("solid", start_color="2F4F4F")
    row_label_fill = PatternFill("solid", start_color="EEEEEE")
    curtain_fill = PatternFill("solid", start_color="FFF8DC")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_thick = Side(style="medium", color="666666")

    n_rows, n_cols = pivot.shape
    title = (
        f"Z-score (per-atom normalised) — colors + curtain  ×  ambiances  @  '{SPACE}'"
    )
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols + 1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=3, column=1, value="atom \\ ambiance").font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=3, column=1).border = border

    for j, col in enumerate(pivot.columns, start=2):
        c = ws.cell(row=3, column=j, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, idx in enumerate(pivot.index, start=4):
        is_curtain = idx == "curtain"
        label = ws.cell(row=i, column=1, value=idx)
        label.font = curtain_font if is_curtain else row_label_font
        label.fill = curtain_fill if is_curtain else row_label_fill
        label.alignment = Alignment(horizontal="left", vertical="center")
        label.border = border

        for j, col in enumerate(pivot.columns, start=2):
            v = pivot.iloc[i - 4, j - 2]
            cell = ws.cell(row=i, column=j, value=float(v) if pd.notna(v) else None)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "0.000;-0.000;-"
            cell.border = border
            if is_curtain:
                cell.fill = curtain_fill

    curtain_excel_row = 3 + len(pivot.index)
    if pivot.index[-1] == "curtain":
        for j in range(1, n_cols + 2):
            cell = ws.cell(row=curtain_excel_row, column=j)
            existing = cell.border
            cell.border = Border(
                left=existing.left, right=existing.right,
                top=top_thick, bottom=existing.bottom,
            )

    last_col_letter = get_column_letter(n_cols + 1)
    data_range = f"B4:{last_col_letter}{3 + n_rows}"
    rule = ColorScaleRule(
        start_type="num", start_value=-2.5, start_color="2C7BB6",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=2.5, end_color="D7191C",
    )
    ws.conditional_formatting.add(data_range, rule)

    ws.column_dimensions["A"].width = 22
    for j in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 32

    ws.freeze_panes = "B4"
    ws.sheet_view.zoomScale = 110

    legend_row = 4 + n_rows + 1
    ws.cell(row=legend_row, column=1, value="Legend").font = title_font
    ws.cell(row=legend_row + 1, column=1,
            value="z-score = (weighted_score − atom mean) / atom std, computed across all 400 (space, ambiance) pairs."
            ).font = body_font
    ws.cell(row=legend_row + 2, column=1,
            value="Positive (red) → atom is more aligned with this ambiance than its own average; negative (blue) → less."
            ).font = body_font
    ws.cell(row=legend_row + 3, column=1,
            value="Each row is normalised independently, so the table answers 'where does this atom shine?' — not 'who wins this column?'."
            ).font = body_font
    for k in range(4):
        ws.merge_cells(start_row=legend_row + k, start_column=1,
                       end_row=legend_row + k, end_column=min(n_cols + 1, 10))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[io] Saved styled XLSX → {path}")


if __name__ == "__main__":
    sys.exit(main())
